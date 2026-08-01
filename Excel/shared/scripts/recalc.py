#!/usr/bin/env python3
"""
Recalculate every formula in a workbook using headless LibreOffice, so that
cached values exist for anything downstream (pandas, data_only=True reads,
previews). openpyxl never computes formulas itself -- it just writes the
formula string with no cached result -- so this step is mandatory any time a
workbook you produced or edited contains formulas.

Usage:
    python recalc.py path/to/workbook.xlsx [timeout_seconds]

Behavior:
    - Rewrites the file in place with LibreOffice-computed cached values.
    - Prints a JSON report to stdout: status, total_formulas, total_errors,
      and an error_summary keyed by error type (#REF!, #NAME?, #DIV/0!, etc.)
      with up to 100 cell locations per type.
    - Exits non-zero ONLY when nothing could be recalculated at all (e.g.
      LibreOffice missing, file locked, corrupt workbook). A workbook that
      recalculates cleanly but contains formula errors still exits 0 with
      status "errors_found" -- always check the JSON, not just the exit code.
"""
import json
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print(json.dumps({"error": "openpyxl not installed"}))
    sys.exit(1)

ERROR_TOKENS = ["#REF!", "#NAME?", "#DIV/0!", "#VALUE!", "#N/A", "#NULL!", "#NUM!", "#SPILL!"]


def find_soffice():
    for candidate in ("soffice", "libreoffice"):
        path = shutil.which(candidate)
        if path:
            return path
    return None


def main():
    if len(sys.argv) < 2:
        print(json.dumps({"error": "usage: recalc.py <workbook.xlsx> [timeout_seconds]"}))
        sys.exit(1)

    target = Path(sys.argv[1]).resolve()
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30

    if not target.exists():
        print(json.dumps({"error": f"file not found: {target}"}))
        sys.exit(1)

    soffice = find_soffice()
    if not soffice:
        print(json.dumps({"error": "LibreOffice (soffice) not found on PATH"}))
        sys.exit(1)

    # Guard against external-workbook links, which LibreOffice cannot
    # re-resolve on this machine and will silently blow away.
    wb_check = openpyxl.load_workbook(target, data_only=False)
    has_external_link_formula = False
    for ws in wb_check.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("=[") or (
                    isinstance(cell.value, str) and "]" in cell.value and cell.value.startswith("=") and cell.value[1:2] == "["
                ):
                    has_external_link_formula = True
    if has_external_link_formula and "--force" not in sys.argv:
        print(json.dumps({
            "error": "external workbook links detected ([1]Sheet!...); recalculating will strip "
                     "their cached values because the linked file is not on disk here. Copy the "
                     "needed values out first, or re-run with --force to accept the loss."
        }))
        sys.exit(1)

    with tempfile.TemporaryDirectory() as tmpdir:
        cmd = [
            soffice, "--headless", "--norestore", "--convert-to", "xlsx",
            "--outdir", tmpdir, str(target),
        ]
        try:
            subprocess.run(cmd, check=True, timeout=timeout, capture_output=True)
        except subprocess.TimeoutExpired:
            print(json.dumps({"error": f"LibreOffice timed out after {timeout}s"}))
            sys.exit(1)
        except subprocess.CalledProcessError as e:
            print(json.dumps({"error": f"LibreOffice failed: {e.stderr.decode(errors='ignore')[:500]}"}))
            sys.exit(1)

        converted = Path(tmpdir) / target.name
        if not converted.exists():
            # LibreOffice sometimes keeps the original extension casing/name differently
            candidates = list(Path(tmpdir).glob("*.xlsx"))
            if not candidates:
                print(json.dumps({"error": "LibreOffice did not produce an output file"}))
                sys.exit(1)
            converted = candidates[0]

        shutil.copy(converted, target)

    # Now scan the recalculated file for errors and formula counts.
    wb = openpyxl.load_workbook(target, data_only=True)
    total_formulas = 0
    total_errors = 0
    error_summary = {}

    wb_formulas = openpyxl.load_workbook(target, data_only=False)
    for sheet_name in wb.sheetnames:
        ws_values = wb[sheet_name]
        ws_formulas = wb_formulas[sheet_name]
        for row_v, row_f in zip(ws_values.iter_rows(), ws_formulas.iter_rows()):
            for cell_v, cell_f in zip(row_v, row_f):
                is_formula = isinstance(cell_f.value, str) and cell_f.value.startswith("=")
                if is_formula:
                    total_formulas += 1
                val = cell_v.value
                if isinstance(val, str) and val in ERROR_TOKENS:
                    total_errors += 1
                    bucket = error_summary.setdefault(val, {"count": 0, "locations": [], "locations_truncated": 0})
                    bucket["count"] += 1
                    loc = f"{sheet_name}!{cell_v.coordinate}"
                    if len(bucket["locations"]) < 100:
                        bucket["locations"].append(loc)
                    else:
                        bucket["locations_truncated"] += 1

    status = "errors_found" if total_errors else "success"
    print(json.dumps({
        "status": status,
        "total_formulas": total_formulas,
        "total_errors": total_errors,
        "error_summary": error_summary,
    }, indent=2))
    sys.exit(0)


if __name__ == "__main__":
    main()
